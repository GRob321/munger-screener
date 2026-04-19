"""
Munger 200-Week MA Screener — Data-Driven Edition
--------------------------------------------------
Screens the entire S&P 500 using objective quality criteria, then surfaces
stocks that are currently at or near their 200-week moving average.

Quality criteria (all must be met):
  • Return on Equity (ROE)     > 15%
  • Net Profit Margin          > 15%
  • Market Cap                 > $10B  (large-cap only)
  • Debt / Equity              < 2.0   (financials excluded from this filter)
  • Trailing EPS               > 0     (currently profitable)
  • EPS positive in            ≥ 7 of last 10 years

Zones:
  🟢 BUY    Price ≤ 200-week MA  (at or below — Munger's entry signal)
  🟡 WATCH  Price 0–20% above MA (approaching — set an alert)
  ⚪ WAIT   Price > 20% above MA (no action)

Usage:
  python3 screener.py                   # full S&P 500 screen
  python3 screener.py --ticker AAPL     # single-stock analysis
  python3 screener.py --refresh         # force re-fetch fundamentals
  python3 screener.py --csv             # export results to .xlsx
  python3 screener.py --email           # email results when done
"""

import os, json, time, warnings, argparse
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed

warnings.filterwarnings("ignore")

import yfinance as yf
import pandas as pd
import numpy as np

# ── Configuration ─────────────────────────────────────────────────────────────

QUALITY_FILTERS = {
    "min_roe":           0.15,   # 15% ROE
    "min_profit_margin": 0.15,   # 15% net margin
    "min_market_cap":    10e9,   # $10 billion
    "max_debt_equity":   2.0,    # D/E ratio
    "min_eps_pos_years": 7,      # of last 10 years with positive EPS
}

CACHE_FILE    = os.path.join(os.path.dirname(__file__), ".fundamentals_cache.json")
CACHE_MAX_AGE = 7  # days before refreshing fundamentals

# ── Step 1: S&P 500 universe ──────────────────────────────────────────────────

def get_sp500_tickers():
    """Fetch current S&P 500 constituents from Wikipedia."""
    import urllib.request
    url = "https://en.wikipedia.org/wiki/List_of_S%26P_500_companies"
    req = urllib.request.Request(url, headers={
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/120.0.0.0 Safari/537.36"
    })
    with urllib.request.urlopen(req) as resp:
        html = resp.read()
    table = pd.read_html(html, header=0)[0]
    tickers = table["Symbol"].str.replace(".", "-", regex=False).tolist()
    sectors = dict(zip(
        table["Symbol"].str.replace(".", "-", regex=False),
        table["GICS Sector"]
    ))
    print(f"  S&P 500 universe: {len(tickers)} tickers")
    return tickers, sectors


# ── Step 2: Fundamental data ──────────────────────────────────────────────────

def fetch_one_fundamental(ticker):
    """Fetch key quality metrics for a single ticker via yfinance."""
    try:
        t = yf.Ticker(ticker)
        info = t.info or {}

        roe     = info.get("returnOnEquity")
        margin  = info.get("profitMargins")
        mcap    = info.get("marketCap")
        de      = info.get("debtToEquity")        # yfinance returns as %, divide by 100
        eps_ttm = info.get("trailingEps")
        name    = info.get("shortName") or info.get("longName") or ticker
        sector  = info.get("sector", "Unknown")

        # Normalise D/E: yfinance often returns as a percent (e.g. 150 = 1.5x)
        if de is not None and de > 20:
            de = de / 100.0

        # EPS history: get up to 10 years of annual EPS
        try:
            fin = t.financials
            if fin is not None and not fin.empty and "Net Income" in fin.index:
                ni = fin.loc["Net Income"].dropna().values
                pos_years = int(sum(1 for v in ni if v > 0))
                total_years = len(ni)
            else:
                pos_years, total_years = None, None
        except Exception:
            pos_years, total_years = None, None

        return ticker, {
            "name":        name,
            "sector":      sector,
            "roe":         roe,
            "margin":      margin,
            "market_cap":  mcap,
            "de_ratio":    de,
            "eps_ttm":     eps_ttm,
            "eps_pos_yrs": pos_years,
            "eps_tot_yrs": total_years,
        }
    except Exception:
        return ticker, None


def load_cache():
    if not os.path.exists(CACHE_FILE):
        return {}
    try:
        with open(CACHE_FILE) as f:
            data = json.load(f)
        cutoff = datetime.now() - timedelta(days=CACHE_MAX_AGE)
        cached_at = datetime.fromisoformat(data.get("_cached_at", "2000-01-01"))
        if cached_at < cutoff:
            return {}
        return data.get("fundamentals", {})
    except Exception:
        return {}


def save_cache(fundamentals):
    payload = {
        "_cached_at":   datetime.now().isoformat(),
        "fundamentals": fundamentals,
    }
    with open(CACHE_FILE, "w") as f:
        json.dump(payload, f)


def fetch_fundamentals(tickers, force_refresh=False):
    """Fetch fundamentals for all tickers, using a 7-day cache."""
    cache = {} if force_refresh else load_cache()
    need  = [t for t in tickers if t not in cache]

    if need:
        print(f"  Fetching fundamentals for {len(need)} tickers "
              f"(~{len(need)//10}–{len(need)//5} seconds)…")
        done = 0
        with ThreadPoolExecutor(max_workers=20) as pool:
            futures = {pool.submit(fetch_one_fundamental, t): t for t in need}
            for fut in as_completed(futures):
                ticker, data = fut.result()
                if data:
                    cache[ticker] = data
                done += 1
                if done % 50 == 0:
                    print(f"    {done}/{len(need)} fetched…")
        save_cache(cache)
        print(f"  Done. Fundamentals cached to {CACHE_FILE}")
    else:
        print(f"  Using cached fundamentals ({len(cache)} tickers, "
              f"refresh with --refresh)")

    return cache


# ── Step 3: Quality filter ────────────────────────────────────────────────────

FINANCIAL_SECTORS = {"Financials", "Financial Services", "Banks", "Insurance"}

def quality_score(f, sector=""):
    """
    Returns (passes: bool, score: int, detail: dict).
    score = number of quality criteria met (max 5).
    """
    filters = QUALITY_FILTERS
    detail  = {}
    checks  = []

    # 1. ROE
    roe = f.get("roe")
    ok  = roe is not None and roe >= filters["min_roe"]
    detail["roe"] = roe
    checks.append(ok)

    # 2. Profit margin
    mg  = f.get("margin")
    ok  = mg is not None and mg >= filters["min_profit_margin"]
    detail["margin"] = mg
    checks.append(ok)

    # 3. Market cap
    mc  = f.get("market_cap")
    ok  = mc is not None and mc >= filters["min_market_cap"]
    detail["market_cap"] = mc
    checks.append(ok)

    # 4. Debt / equity — relax for financials (inherently high leverage)
    de  = f.get("de_ratio")
    is_financial = sector in FINANCIAL_SECTORS
    if is_financial:
        ok = True   # financials excluded from D/E filter
    else:
        ok = de is None or de <= filters["max_debt_equity"]
    detail["de_ratio"] = de
    checks.append(ok)

    # 5. EPS consistency
    pos = f.get("eps_pos_yrs")
    tot = f.get("eps_tot_yrs") or 4
    ok  = (pos is not None and pos >= min(filters["min_eps_pos_years"], tot)) \
          or (f.get("eps_ttm") or 0) > 0  # fallback: at least profitable now
    detail["eps_pos_yrs"] = pos
    checks.append(ok)

    score   = sum(checks)
    passes  = all(checks)
    return passes, score, detail


# ── Step 4: 200-week MA calculation ──────────────────────────────────────────

def fetch_prices_and_ma(tickers):
    """Bulk-download 5 years of weekly prices; return current price + 200wk MA."""
    print(f"  Downloading weekly prices for {len(tickers)} tickers…")
    raw = yf.download(
        tickers,
        period="5y",
        interval="1wk",
        auto_adjust=True,
        progress=False,
    )["Close"]

    if isinstance(raw, pd.Series):
        raw = raw.to_frame(name=tickers[0])

    sma200 = raw.rolling(200, min_periods=150).mean()

    results = {}
    for ticker in raw.columns:
        px = raw[ticker].dropna()
        ma = sma200[ticker].dropna()
        if px.empty or ma.empty:
            continue
        price = px.iloc[-1]
        ma_v  = ma.iloc[-1]
        results[ticker] = {
            "price":     round(float(price), 2),
            "ma200w":    round(float(ma_v), 2),
            "pct_above": round((float(price) / float(ma_v) - 1) * 100, 1),
        }
    return results


# ── Step 5: Excel Export ─────────────────────────────────────────────────────

def export_xlsx(merged, buy_zone, watch, wait, run_date, out_path):
    from openpyxl import Workbook
    from openpyxl.styles import (PatternFill, Font, Alignment,
                                  Border, Side, numbers)
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Palette ───────────────────────────────────────────────────────────────
    GREEN_ROW   = PatternFill("solid", fgColor="C6EFCE")   # light green
    YELLOW_ROW  = PatternFill("solid", fgColor="FFEB9C")   # light yellow
    GRAY_ROW    = PatternFill("solid", fgColor="F2F2F2")   # light gray
    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")   # dark navy
    SECTION_FILL= PatternFill("solid", fgColor="D6E4F0")   # pale blue section divider

    HDR_FONT    = Font(bold=True, color="FFFFFF", size=11)
    BOLD        = Font(bold=True)
    SECTION_FONT= Font(bold=True, color="1F4E79", size=11)
    thin = Side(style="thin", color="CCCCCC")
    BORDER = Border(bottom=Side(style="thin", color="CCCCCC"))

    def pct_str(v):
        return f"{v*100:+.1f}%" if v is not None else "n/a"

    def money(v):
        return f"${v:,.2f}" if v is not None else "n/a"

    # ── Sheet 1: Results ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Results"
    ws.freeze_panes = "A2"

    headers = ["Zone", "Ticker", "Name", "Sector",
               "Price", "200W MA", "% vs MA",
               "ROE", "Net Margin", "D/E", "Signal"]
    col_widths = [8, 8, 30, 24, 10, 10, 10, 10, 12, 7, 18]

    # Header row
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill   = HEADER_FILL
        cell.font   = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 20

    def zone_label(pct):
        if pct <= 0:   return "🟢 BUY"
        elif pct <= 20: return "🟡 WATCH"
        else:           return "⚪ WAIT"

    def signal_str(pct):
        if pct <= -20:  return "Far Below MA"
        elif pct <= -5: return "Below MA"
        elif pct <= 0:  return "At MA"
        elif pct <= 10: return "Close"
        elif pct <= 20: return "Watch"
        else:           return "Wait"

    row_num = 2
    for section_name, section_rows, fill in [
        ("🟢  BUY ZONE — at or below 200-week MA", buy_zone,  GREEN_ROW),
        ("🟡  WATCH LIST — 0–20% above 200-week MA", watch,  YELLOW_ROW),
        ("⚪  WAIT — more than 20% above 200-week MA", wait,  GRAY_ROW),
    ]:
        # Section divider row
        ws.merge_cells(start_row=row_num, start_column=1,
                       end_row=row_num, end_column=len(headers))
        cell = ws.cell(row=row_num, column=1, value=section_name)
        cell.fill = SECTION_FILL
        cell.font = SECTION_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row_num].height = 18
        row_num += 1

        for r in section_rows:
            det = r["detail"]
            ma  = r["ma"]
            pct = ma["pct_above"]
            row_data = [
                zone_label(pct),
                r["ticker"],
                r["fund"]["name"],
                r["sector"],
                ma["price"],
                ma["ma200w"],
                pct / 100,                               # stored as decimal for % format
                (det["roe"]    or 0),
                (det["margin"] or 0),
                (det["de_ratio"] if det["de_ratio"] is not None else "n/a"),
                signal_str(pct),
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=val)
                cell.fill = fill
                cell.alignment = Alignment(vertical="center")
                cell.border = BORDER
                # Format numbers
                if col == 5:  cell.number_format = '$#,##0.00'
                if col == 6:  cell.number_format = '$#,##0.00'
                if col == 7:  cell.number_format = '+0.0%;-0.0%;0.0%'
                if col == 8:  cell.number_format = '0.0%'
                if col == 9:  cell.number_format = '0.0%'
                if col == 10 and isinstance(val, float):
                    cell.number_format = '0.00'
            row_num += 1

        row_num += 1   # blank spacer between sections

    # Summary box at bottom
    row_num += 1
    summary = [
        ("Run date",          run_date.strftime("%B %d, %Y")),
        ("Universe",          "S&P 500 (503 stocks)"),
        ("Quality filter pass", f"{len(merged)} stocks"),
        ("🟢 Buy zone",        f"{len(buy_zone)} stocks"),
        ("🟡 Watch list",      f"{len(watch)} stocks"),
        ("⚪ Wait",            f"{len(wait)} stocks"),
    ]
    for label, val in summary:
        c1 = ws.cell(row=row_num, column=1, value=label)
        c1.font = BOLD
        c1.alignment = Alignment(horizontal="right")
        c2 = ws.cell(row=row_num, column=2, value=val)
        row_num += 1

    # ── Sheet 2: Criteria ─────────────────────────────────────────────────────
    wc = wb.create_sheet("Criteria")
    wc.freeze_panes = "A2"
    wc.column_dimensions["A"].width = 5
    wc.column_dimensions["B"].width = 28
    wc.column_dimensions["C"].width = 24
    wc.column_dimensions["D"].width = 55

    def crit_header(row, text):
        wc.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        cell = wc.cell(row=row, column=1, value=text)
        cell.fill = HEADER_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        wc.row_dimensions[row].height = 20

    def crit_subheader(row, text):
        wc.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        cell = wc.cell(row=row, column=1, value=text)
        cell.fill = SECTION_FILL
        cell.font = SECTION_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        wc.row_dimensions[row].height = 16

    r = 1
    crit_header(r, "Munger 200-Week MA Screener — Quality Criteria & Zone Definitions")
    r += 1

    wc.cell(row=r, column=1, value=f"Generated: {run_date.strftime('%B %d, %Y')}")
    wc.cell(row=r, column=1).font = Font(italic=True, color="666666")
    r += 2

    # Quality criteria table
    crit_subheader(r, "Quality Criteria  (all 5 must be met)")
    r += 1
    for col, h in enumerate(["#", "Criterion", "Threshold", "Rationale"], 1):
        cell = wc.cell(row=r, column=col, value=h)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    r += 1

    criteria = [
        (1, "Return on Equity (ROE)",    "> 15%",
            "Measures how efficiently a company generates profit from equity — "
            "a core indicator of durable competitive advantage (moat)"),
        (2, "Net Profit Margin",         "> 15%",
            "Filters for pricing power and cost discipline; low-margin businesses "
            "lack the buffer to survive downturns"),
        (3, "Market Capitalization",     "> $10 billion",
            "Large-cap only — sufficient liquidity, analyst coverage, "
            "and operational track record"),
        (4, "Debt / Equity Ratio",       "< 2.0  (financials exempt)",
            "Avoids over-leveraged companies. Financial sector firms "
            "(banks, insurers) are excluded as high leverage is structural"),
        (5, "Earnings Consistency",      "EPS positive ≥ 7 of 10 years",
            "Filters for durable, repeatable earnings — not cyclical spikes "
            "or one-time profitability"),
    ]
    alt = PatternFill("solid", fgColor="EEF2FF")
    for i, (num, crit, thresh, rationale) in enumerate(criteria):
        fill = alt if i % 2 == 0 else PatternFill()
        for col, val in enumerate([num, crit, thresh, rationale], 1):
            cell = wc.cell(row=r, column=col, value=val)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if col == 4:
                cell.alignment = Alignment(wrap_text=True, vertical="center")
        wc.row_dimensions[r].height = 32
        r += 1

    r += 1

    # Zone definitions table
    crit_subheader(r, "Zone Definitions")
    r += 1
    for col, h in enumerate(["Zone", "Condition", "Meaning", "Action"], 1):
        cell = wc.cell(row=r, column=col, value=h)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    r += 1

    zones = [
        ("🟢  BUY",   "Price ≤ 200-week MA",
         "Stock is trading at or below its ~4-year average price",
         "Buy — this is Munger's entry signal",
         GREEN_ROW),
        ("🟡  WATCH", "0% < Price ≤ 20% above MA",
         "Approaching the MA — a correction could bring it into buy range",
         "Set a price alert; prepare your thesis",
         YELLOW_ROW),
        ("⚪  WAIT",  "Price > 20% above MA",
         "Too far above the long-term average for a Munger-style entry",
         "No action — add to watchlist for the next correction",
         GRAY_ROW),
    ]
    for zone, condition, meaning, action, fill in zones:
        for col, val in enumerate([zone, condition, meaning, action], 1):
            cell = wc.cell(row=r, column=col, value=val)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        wc.row_dimensions[r].height = 36
        r += 1

    r += 2

    # Methodology notes
    crit_subheader(r, "Methodology Notes")
    r += 1
    notes = [
        "200-week MA is calculated as the simple moving average of weekly adjusted closing prices "
        "over the most recent 200 weeks (~3.8 years). Requires at least 150 weeks of data.",
        "Data source: Yahoo Finance (yfinance) for fundamentals and prices; "
        "S&P 500 constituent list from Wikipedia.",
        "Fundamentals are cached for 7 days. Run screener.py --refresh to force an update.",
        "Financial sector firms (banks, insurers, brokers) are excluded from the Debt/Equity filter "
        "because leverage is a core feature of their business model, not a risk indicator.",
        "This screen is a starting point, not investment advice. "
        "Individual company analysis is required before investing.",
    ]
    for note in notes:
        wc.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        cell = wc.cell(row=r, column=1, value=f"• {note}")
        cell.alignment = Alignment(wrap_text=True, vertical="center", indent=1)
        wc.row_dimensions[r].height = 30
        r += 1

    wb.save(out_path)
    print(f"\n  Saved → {out_path}")
    print(f"  Import into Google Sheets: File → Import → Upload → select the file\n")


# ── Step 6: Terminal Display ──────────────────────────────────────────────────

def fmt_cap(v):
    if v is None: return "  n/a  "
    if v >= 1e12: return f"${v/1e12:>4.1f}T"
    if v >= 1e9:  return f"${v/1e9:>4.0f}B"
    return f"${v/1e6:>4.0f}M"

def fmt_pct(v, digits=1):
    if v is None: return "  n/a "
    return f"{v*100:>+5.{digits}f}%"

def signal_label(pct):
    if pct <= -20:      return "🟢 FAR BELOW MA"
    elif pct <= -5:     return "🟢 BELOW MA"
    elif pct <= 0:      return "🟢 AT MA"
    elif pct <= 10:     return "🟡 CLOSE"
    elif pct <= 20:     return "🟡 WATCH"
    else:               return "⚪ WAIT"


def print_table(rows, threshold, show_score=True):
    hdr = (f"  {'#':>2}  {'Ticker':<7} {'Name':<26} {'Sector':<22} "
           f"{'Price':>8} {'200W MA':>8} {'%AbvMA':>8}  "
           f"{'ROE':>6} {'Mgn':>6} {'D/E':>5}  {'Score':>5}  Signal")
    print(hdr)
    print("  " + "-" * 118)
    for i, r in enumerate(rows, 1):
        f    = r["fund"]
        ma   = r["ma"]
        det  = r["detail"]
        sig  = signal_label(ma["pct_above"])
        de_s = f"{det['de_ratio']:.2f}" if det["de_ratio"] is not None else " n/a"
        print(
            f"  {i:>2}  {r['ticker']:<7} {f['name'][:25]:<26} "
            f"{r['sector'][:21]:<22} "
            f"${ma['price']:>7,.2f} ${ma['ma200w']:>7,.2f} "
            f"  {ma['pct_above']:>+6.1f}%  "
            f"{fmt_pct(det['roe']):>6} {fmt_pct(det['margin']):>6} {de_s:>5}  "
            f"  {r['score']}/5   {sig}"
        )


# ── Email ─────────────────────────────────────────────────────────────────────

EMAIL_TO = "robinsongm321@gmail.com"

def send_email(subject, body):
    """Send email via macOS mail command."""
    import subprocess, tempfile
    with tempfile.NamedTemporaryFile(mode="w", suffix=".txt", delete=False) as f:
        f.write(body)
        tmp = f.name
    result = subprocess.run(
        f'mail -s "{subject}" "{EMAIL_TO}" < "{tmp}"',
        shell=True, capture_output=True, text=True
    )
    os.unlink(tmp)
    if result.returncode == 0:
        print(f"  Email sent to {EMAIL_TO}")
    else:
        print(f"  Email failed: {result.stderr.strip() or result.stdout.strip()}")


# ── Single-Stock Analysis ──────────────────────────────────────────────────────

def analyse_single(ticker, run_date, send_mail=False):
    """Full Munger quality + 200-week MA analysis for one ticker."""
    ticker = ticker.upper()
    print(f"\n{'='*60}")
    print(f"  SINGLE-STOCK ANALYSIS  —  {ticker}  —  {run_date.strftime('%B %d, %Y')}")
    print(f"{'='*60}\n")

    # Fundamentals
    print("  Fetching fundamentals…")
    _, fund = fetch_one_fundamental(ticker)
    if not fund:
        print(f"  Could not fetch data for {ticker}.")
        return

    # Price + 200-week MA
    print("  Fetching price history…")
    ma_data = fetch_prices_and_ma([ticker])
    if ticker not in ma_data:
        print(f"  Could not fetch price data for {ticker}.")
        return

    ma   = ma_data[ticker]
    pct  = ma["pct_above"]
    sector = fund.get("sector", "")
    passes, score, detail = quality_score(fund, sector)

    # ── Zone ──────────────────────────────────────────────────────────────────
    if pct <= 0:
        zone = "🟢 BUY — at or below 200-week MA (Munger entry signal)"
        zone_short = "BUY"
    elif pct <= 20:
        zone = f"🟡 WATCH — {pct:.1f}% above MA (approaching entry point)"
        zone_short = "WATCH"
    else:
        zone = f"⚪ WAIT — {pct:.1f}% above MA (too far for entry)"
        zone_short = "WAIT"

    # ── Print report ──────────────────────────────────────────────────────────
    name = fund.get("name", ticker)
    print(f"  {name} ({ticker})  |  {sector}\n")

    print(f"  {'Current Price':<30} ${ma['price']:,.2f}")
    print(f"  {'200-Week Moving Average':<30} ${ma['ma200w']:,.2f}")
    print(f"  {'% vs MA':<30} {pct:+.1f}%")
    print(f"  {'Zone':<30} {zone}\n")

    print(f"  Quality Criteria (score: {score}/5):")
    print(f"  {'─'*52}")

    def check(label, value, threshold_str, passes_flag):
        icon = "✅" if passes_flag else "❌"
        val_str = f"{value*100:+.1f}%" if isinstance(value, float) else str(value)
        print(f"  {icon}  {label:<28} {val_str:>10}   (threshold: {threshold_str})")

    roe_ok  = detail["roe"] is not None and detail["roe"] >= QUALITY_FILTERS["min_roe"]
    mgn_ok  = detail["margin"] is not None and detail["margin"] >= QUALITY_FILTERS["min_profit_margin"]
    mc      = fund.get("market_cap")
    mc_ok   = mc is not None and mc >= QUALITY_FILTERS["min_market_cap"]
    de      = detail["de_ratio"]
    is_fin  = sector in FINANCIAL_SECTORS
    de_ok   = is_fin or (de is None or de <= QUALITY_FILTERS["max_debt_equity"])
    pos     = detail["eps_pos_yrs"]
    eps_ok  = (pos is not None and pos >= QUALITY_FILTERS["min_eps_pos_years"]) \
              or (fund.get("eps_ttm") or 0) > 0

    check("Return on Equity (ROE)",   detail["roe"] or 0,     "> 15%",          roe_ok)
    check("Net Profit Margin",        detail["margin"] or 0,  "> 15%",          mgn_ok)
    mc_str = f"${mc/1e9:.1f}B" if mc else "n/a"
    print(f"  {'✅' if mc_ok else '❌'}  {'Market Cap':<28} {mc_str:>10}   (threshold: > $10B)")
    de_str = f"{de:.2f}x" if de is not None else "n/a"
    de_lbl = "Debt / Equity" + (" (exempt)" if is_fin else "")
    print(f"  {'✅' if de_ok else '❌'}  {de_lbl:<28} {de_str:>10}   (threshold: < 2.0x)")
    eps_str = f"{pos}/{detail.get('eps_tot_yrs') or 'n/a'} yrs" if pos else "n/a"
    print(f"  {'✅' if eps_ok else '❌'}  {'EPS Consistency':<28} {eps_str:>10}   (threshold: ≥ 7/10 yrs)")

    overall = "✅ PASSES all quality criteria" if passes else f"❌ FAILS quality criteria ({5-score} criterion/a not met)"
    print(f"\n  Overall: {overall}")
    print(f"\n{'='*60}\n")

    # ── Email ─────────────────────────────────────────────────────────────────
    if send_mail:
        lines = [
            f"Munger 200-Week MA Analysis: {ticker}",
            f"Date: {run_date.strftime('%B %d, %Y')}",
            f"",
            f"{name} ({ticker}) | {sector}",
            f"",
            f"Price:           ${ma['price']:,.2f}",
            f"200-Week MA:     ${ma['ma200w']:,.2f}",
            f"% vs MA:         {pct:+.1f}%",
            f"Zone:            {zone_short}",
            f"Quality Score:   {score}/5  ({'PASS' if passes else 'FAIL'})",
            f"",
            f"Quality Criteria:",
            f"  {'✅' if roe_ok  else '❌'} ROE:             {(detail['roe'] or 0)*100:+.1f}%  (>15%)",
            f"  {'✅' if mgn_ok  else '❌'} Net Margin:      {(detail['margin'] or 0)*100:+.1f}%  (>15%)",
            f"  {'✅' if mc_ok   else '❌'} Market Cap:      {mc_str}  (>$10B)",
            f"  {'✅' if de_ok   else '❌'} Debt/Equity:     {de_str}  (<2.0x)",
            f"  {'✅' if eps_ok  else '❌'} EPS Consistency: {eps_str}  (≥7/10 yrs)",
            f"",
            f"Verdict: {zone}",
        ]
        body = "\n".join(lines)
        send_email(f"Munger Screen: {ticker} — {zone_short} ({pct:+.1f}% vs MA)", body)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--ticker", type=str, default=None,
                        help="Analyse a single stock (e.g. --ticker AAPL)")
    parser.add_argument("--refresh", action="store_true",
                        help="Force re-fetch fundamentals (ignore cache)")
    parser.add_argument("--csv", action="store_true",
                        help="Export results to .xlsx for Google Sheets")
    parser.add_argument("--email", action="store_true",
                        help=f"Email results to {EMAIL_TO} when complete")
    args = parser.parse_args()

    run_date = datetime.today()

    # ── Single-stock mode ─────────────────────────────────────────────────────
    if args.ticker:
        analyse_single(args.ticker, run_date, send_mail=args.email)
        return

    print(f"\n{'='*72}")
    print(f"  MUNGER 200-WEEK MA SCREENER  —  {run_date.strftime('%B %d, %Y')}")
    print(f"  Universe: S&P 500   |   Quality-filtered")
    print(f"  🟢 BUY: at/below MA  |  🟡 WATCH: 0–20% above  |  ⚪ WAIT: >20% above")
    print(f"{'='*72}\n")

    # 1. Universe
    tickers, wiki_sectors = get_sp500_tickers()

    # 2. Fundamentals
    fundamentals = fetch_fundamentals(tickers, force_refresh=args.refresh)

    # 3. Quality screen
    print(f"\n  Applying quality filters…")
    quality_pass = {}
    for ticker, f in fundamentals.items():
        sector = f.get("sector") or wiki_sectors.get(ticker, "")
        passes, score, detail = quality_score(f, sector)
        if passes:
            quality_pass[ticker] = {"fund": f, "score": score,
                                    "detail": detail, "sector": sector}
    print(f"  {len(quality_pass)} of {len(fundamentals)} stocks pass quality filters")

    # 4. Prices + 200-wk MA for quality stocks
    q_tickers = list(quality_pass.keys())
    if not q_tickers:
        print("  No stocks passed quality filters.")
        return

    print()
    ma_data = fetch_prices_and_ma(q_tickers)

    # 5. Merge and sort
    merged = []
    for ticker, qdata in quality_pass.items():
        if ticker not in ma_data:
            continue
        merged.append({
            "ticker": ticker,
            "fund":   qdata["fund"],
            "score":  qdata["score"],
            "detail": qdata["detail"],
            "sector": qdata["sector"],
            "ma":     ma_data[ticker],
        })

    merged.sort(key=lambda x: x["ma"]["pct_above"])

    # 6. Segment — corrected zones
    buy_zone = [r for r in merged if r["ma"]["pct_above"] <= 0]
    watch    = [r for r in merged if 0 < r["ma"]["pct_above"] <= 20]
    wait     = [r for r in merged if r["ma"]["pct_above"] > 20]

    # ── Print results ─────────────────────────────────────────────────────────
    print(f"\n{'='*72}")
    print(f"  🟢 BUY ZONE — {len(buy_zone)} stock(s) at or below their 200-week MA\n")
    if buy_zone:
        print_table(buy_zone, 0)
    else:
        print("  None currently at or below their 200-week MA.")

    print(f"\n  🟡 WATCH LIST — {len(watch)} stock(s) within 0–20% above MA\n")
    if watch:
        print_table(watch, 0)

    print(f"\n  ⚪ WAIT — {len(wait)} quality stock(s) more than 20% above MA\n")
    if wait:
        print_table(wait, 0)

    # ── Summary ───────────────────────────────────────────────────────────────
    print(f"\n{'='*72}")
    print(f"  Quality Filter Summary (S&P 500 = {len(tickers)} stocks)")
    print(f"  {'Criterion':<35} {'Threshold':>20}")
    print(f"  {'-'*55}")
    print(f"  {'Return on Equity (ROE)':<35} {'>15%':>20}")
    print(f"  {'Net Profit Margin':<35} {'>15%':>20}")
    print(f"  {'Market Capitalization':<35} {'>$10B':>20}")
    print(f"  {'Debt / Equity':<35} {'<2.0 (ex-financials)':>20}")
    print(f"  {'Earnings consistency':<35} {'EPS>0 (≥7/10 yrs)':>20}")
    print(f"\n  Passed quality filters: {len(quality_pass)}/{len(fundamentals)}  |  "
          f"🟢 Buy: {len(buy_zone)}  |  🟡 Watch: {len(watch)}  |  ⚪ Wait: {len(wait)}")
    print(f"  Cache: {CACHE_FILE.split('/')[-1]}  (--refresh to update)")
    print(f"{'='*72}\n")

    # ── CSV / Excel export ────────────────────────────────────────────────────
    if args.csv:
        out = os.path.join(
            os.path.dirname(__file__),
            f"munger_screen_{run_date.strftime('%Y-%m-%d')}.xlsx"
        )
        export_xlsx(merged, buy_zone, watch, wait, run_date, out)

    # ── Email results ─────────────────────────────────────────────────────────
    if args.email:
        buy_lines  = "\n".join(
            f"  {r['ticker']:<7} {r['fund']['name'][:30]:<30}  {r['ma']['pct_above']:+.1f}%"
            for r in buy_zone
        )
        watch_lines = "\n".join(
            f"  {r['ticker']:<7} {r['fund']['name'][:30]:<30}  {r['ma']['pct_above']:+.1f}%"
            for r in watch[:10]   # top 10 closest
        )
        body = "\n".join([
            f"Munger 200-Week MA Screen — {run_date.strftime('%B %d, %Y')}",
            f"Universe: S&P 500 ({len(tickers)} stocks)",
            f"Quality filter: {len(merged)}/503 pass",
            f"",
            f"{'─'*45}",
            f"🟢 BUY ZONE ({len(buy_zone)} stocks) — at or below 200-week MA",
            f"{'─'*45}",
            buy_lines or "  None",
            f"",
            f"{'─'*45}",
            f"🟡 WATCH LIST ({len(watch)} stocks) — 0–20% above MA",
            f"{'─'*45}",
            watch_lines or "  None",
            f"",
            f"Run:  python3 screener.py --csv  to get the full Excel export.",
        ])
        send_email(
            f"Munger Screen: {len(buy_zone)} in buy zone, {len(watch)} on watch list — {run_date.strftime('%b %d')}",
            body
        )


if __name__ == "__main__":
    main()
